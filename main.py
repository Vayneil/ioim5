import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
from openpyxl import Workbook
import pyswarms as ps

# python implementation of particle swarm optimization (PSO)
# minimizing rastrigin and sphere function

import random
import math  # cos() for Rastrigin
import copy  # array-copying convenience
import sys  # max float

print("Loading excel...")
wb = load_workbook('./Dane.xlsx')
print("Excel loaded.")

rho = []
T = []
e_dot = []

for ws in wb.worksheets:
    rho_test = [ws.cell(i, 2).value for i in range(2, 100002, 10)]
    T.append(ws.cell(1, 7).value)
    e_dot.append(ws.cell(2, 7).value)
    rho.append(rho_test)
print("Data in python.")


# print(len(rho[0]))
# print(len(np.linspace(0.0, 1, 100000)))


def sigma_p(a, T, t_max, e_dot, plot):
    # print(a)
    step = t_max / 1e4
    # e_dot = 1.0
    # t_max = 1.0
    # T = 898.0
    b = 0.25e-9
    D = 30.0
    mu = 45000.0
    Q = 238000.0
    rho_0 = 1e4
    R = 8.314
    Z = e_dot * math.exp(Q / R / T)
    rho = rho_0
    t_0 = 0.0
    t = t_0
    rho_values = []
    t_cr = t_max

    # a1 = 2.1
    # a2 = 176.0
    # a3 = 19.5
    # a4 = 0.000148
    # a5 = 151.0
    # a6 = 0.973
    # a7 = 5.77
    # a8 = 1.0
    # a9 = 0.0
    # a10 = 0.262
    # a11 = 0.0
    # a12 = 0.000605
    # a13 = 0.167

    tau = 1e6 * mu * b ** 2 * 0.5
    l = a[0] * 1e-3 / (Z ** a[12])
    A1 = 1 / b / l
    A2 = a[1] * e_dot ** (-a[8]) * math.exp(-a[2] * 1e3 / R / T)
    A3 = a[3] * 3e10 * tau / D * math.exp(-a[4] * 1e3 / R / T)
    rho_cr = -a[10] * 1e13 + a[11] * 1e13 * Z ** a[9]
    is_critical = False

    def drho_dt(is_critical, t_cr):
        # global is_critical
        # global t_cr
        if not is_critical:
            if rho > rho_cr:
                t_cr = t
                is_critical = True
                t2 = t - t_cr
                round(t2, 3)
                step_number = round(t2 / step)
                X = rho_values[step_number]
            else:
                X = 0
        else:
            t2 = t - t_cr
            round(t2, 3)
            step_number = round(t2 / step)
            X = rho_values[step_number]
        result = A1 * e_dot - A2 * e_dot * rho - A3 * rho ** a[7] * X
        return result, is_critical, t_cr

    while True:
        result, is_critical, t_cr = drho_dt(is_critical, t_cr)
        rho = rho + step * result
        t += step
        rho_values.append(rho)
        if t > t_max:
            break
    # print(len(rho_values))
    if plot:
        plt.plot(np.arange(t_0, t_max, t_max / len(rho_values)), rho_values, label="T = " + str(T))
    return rho_values


def objective(pos):
    particles = pos.shape[0]
    result = np.zeros(particles)
    for k in range(particles):
        sum = 0
        for i in range(len(T)):
            sigma = sigma_p(pos[k], T[i], 1 / e_dot[i], e_dot[i], False)
            for j in range(10000):
                err_squared = (rho[i][j] - sigma[j]) ** 2
                err_relative = err_squared / rho[i][j]
                sum = sum + err_relative
        result[k] = sum

    # print(result)
    return result


swarm_size = 50
dim = 13
options = {'c1': 1, 'c2': 1, 'w': 0.5}
# constraints = (np.array([0.05, 15000, 50, 0.01, 100, 1.5, 0, 0.2, 0.05, 0.1, 0, 0.00001, 0.01]),
#                np.array([10, 22000, 100, 0.09, 150, 2.5, 0, 0.8, 0.25, 0.9, 0, 0.00009, 0.09]))
constraints = (np.array([0.05, 15000, 50, 0.01, 100, 1.5, 1e-12, 0.4519, 0.05, 0.4089, 1e-12, 0.0000419, 0.01]),
               np.array([10, 22000, 100, 0.09, 150, 2.5, 1.01e-12, 0.452, 0.25, 0.409, 1.01e-12, 0.000042, 0.09]))

optimizer = ps.single.GlobalBestPSO(n_particles=swarm_size, dimensions=dim, options=options, bounds=constraints)
cost, joint_vars = optimizer.optimize(objective_func=objective, iters=100)
# results = [2.78482984e-01 1.69558166e+04 7.41077741e+01 6.29923307e-02
#  1.22153107e+02 1.83320733e+00 1.00546722e-12 4.51929246e-01
#  2.06192444e-01 4.08944646e-01 1.00396002e-12 4.19337918e-05
#  8.41070100e-02]


# sigma_p(joint_vars, T[6], 1 / e_dot[6], e_dot[6], True)
# for temp in [1200, 1050, 900]:
#     sigma_p(results, temp, 1 / 10, 10, True)
# plt.show()

# -------fitness functions---------

# # rastrigin function
# def fitness_rastrigin(position):
#   fitnessVal = 0.0
#   for i in range(len(position)):
#     xi = position[i]
#     fitnessVal += (xi * xi) - (10 * math.cos(2 * math.pi * xi)) + 10
#   return fitnessVal
#
# #sphere function
# def fitness_sphere(position):
#     fitnessVal = 0.0
#     for i in range(len(position)):
#         xi = position[i]
#         fitnessVal += (xi*xi)
#     return fitnessVal

# objective function

#
#
# #-------------------------
#
# #particle class
# class Particle:
#   def __init__(self, fitness, dim, minx, maxx, seed):
#     self.rnd = random.Random(seed)
#
#     # initialize position of the particle with 0.0 value
#     self.position = [0.0 for i in range(dim)]
#
#      # initialize velocity of the particle with 0.0 value
#     self.velocity = [0.0 for i in range(dim)]
#
#     # initialize best particle position of the particle with 0.0 value
#     self.best_part_pos = [0.0 for i in range(dim)]
#
#     # loop dim times to calculate random position and velocity
#     # range of position and velocity is [minx, max]
#     for i in range(dim):
#       self.position[i] = ((maxx[i] - minx[i]) *
#         self.rnd.random() + minx[i])
#       if self.position[i] > maxx[i] or self.position[i] < minx[i]:
#           raise Exception("Bad value of position")
#       self.velocity[i] = ((maxx[i] - minx[i]) *
#         self.rnd.random() + minx[i])
#
#     # compute fitness of particle
#     self.fitness = fitness(self.position) # curr fitness
#
#     # initialize best position and fitness of this particle
#     self.best_part_pos = copy.copy(self.position)
#     self.best_part_fitnessVal = self.fitness # best fitness
#
# # particle swarm optimization function
# def pso(fitness, max_iter, n, dim, minx, maxx):
#   # hyper parameters
#   w = 0.729    # inertia
#   c1 = 1.49445 # cognitive (particle)
#   c2 = 1.49445 # social (swarm)
#
#   rnd = random.Random(0)
#
#   print("create n random particles")
#   # create n random particles
#   swarm = [Particle(fitness, dim, minx, maxx, i * 5) for i in range(n)]
#
#   print("compute the value of best_position and best_fitness in swarm")
#   # compute the value of best_position and best_fitness in swarm
#   best_swarm_pos = [0.0 for i in range(dim)]
#   best_swarm_fitnessVal = sys.float_info.max # swarm best
#
#   print("computer best particle of swarm and it's fitness")
#   # computer best particle of swarm and it's fitness
#   for i in range(n): # check each particle
#     if swarm[i].fitness < best_swarm_fitnessVal:
#       best_swarm_fitnessVal = swarm[i].fitness
#       best_swarm_pos = copy.copy(swarm[i].position)
#
#   # main loop of pso
#   Iter = 0
#   print("Iteration:")
#   while Iter < max_iter:
#     print(Iter)
#     # after every 10 iterations
#     # print iteration number and best fitness value so far
#     if Iter % 10 == 0 and Iter > 1:
#       print("Iter = " + str(Iter) + " best fitness = %.3f" % best_swarm_fitnessVal)
#
#     for i in range(n): # process each particle
#
#       # compute new velocity of curr particle
#       for k in range(dim):
#         r1 = rnd.random()    # randomizations
#         r2 = rnd.random()
#
#         swarm[i].velocity[k] = (
#                                  (w * swarm[i].velocity[k]) +
#                                  (c1 * r1 * (swarm[i].best_part_pos[k] - swarm[i].position[k])) +
#                                  (c2 * r2 * (best_swarm_pos[k] -swarm[i].position[k]))
#                                )
#         # print(swarm[i].velocity[k])
#
#         # if velocity[k] is not in [minx, max]
#         # then clip it
#         if swarm[i].velocity[k] < minx[k]:
#           swarm[i].velocity[k] = minx[k]
#         elif swarm[i].velocity[k] > maxx[k]:
#           swarm[i].velocity[k] = maxx[k]
#
#
#       # compute new position using new velocity
#       for k in range(dim):
#         swarm[i].position[k] += swarm[i].velocity[k]
#         if swarm[i].position[k] < minx[k]:
#             swarm[i].position[k] = minx[k]
#         elif swarm[i].position[k] > maxx[k]:
#             swarm[i].position[k] = maxx[k]
#
#
#       # compute fitness of new position
#       swarm[i].fitness = fitness(swarm[i].position)
#
#       # is new position a new best for the particle?
#       if swarm[i].fitness < swarm[i].best_part_fitnessVal:
#         swarm[i].best_part_fitnessVal = swarm[i].fitness
#         swarm[i].best_part_pos = copy.copy(swarm[i].position)
#
#       # is new position a new best overall?
#       if swarm[i].fitness < best_swarm_fitnessVal:
#         best_swarm_fitnessVal = swarm[i].fitness
#         best_swarm_pos = copy.copy(swarm[i].position)
#
#
#     # for-each particle
#     Iter += 1
#   #end_while
#   return best_swarm_pos
# # end pso


# ----------------------------
# Driver code for rastrigin function

# print("\nBegin particle swarm optimization on rastrigin function\n")
# dim = 3
# fitness = fitness_rastrigin


# print("Goal is to minimize Rastrigin's function in " + str(dim) + " variables")
# print("Function has known min = 0.0 at (", end="")
# for i in range(dim-1):
#   print("0, ", end="")
# print("0)")

# num_particles = 50
# max_iter = 100

# print("Setting num_particles = " + str(num_particles))
# print("Setting max_iter    = " + str(max_iter))
# print("\nStarting PSO algorithm\n")


# best_position = pso(fitness, max_iter, num_particles, dim, -10.0, 10.0)

# print("\nPSO completed\n")
# print("\nBest solution found:")
# print(["%.6f"%best_position[k] for k in range(dim)])
# fitnessVal = fitness(best_position)
# print("fitness of best solution = %.6f" % fitnessVal)

# print("\nEnd particle swarm for rastrigin function\n")


# print()
# print()


# # Driver code for Sphere function
# print("\nBegin particle swarm optimization on sphere function\n")
# dim = 3
# fitness = fitness_sphere


# print("Goal is to minimize sphere function in " + str(dim) + " variables")
# print("Function has known min = 0.0 at (", end="")
# for i in range(dim-1):
#   print("0, ", end="")
# print("0)")

# num_particles = 50
# max_iter = 100

# print("Setting num_particles = " + str(num_particles))
# print("Setting max_iter    = " + str(max_iter))
# print("\nStarting PSO algorithm\n")


# best_position = pso(fitness, max_iter, num_particles, dim, -10.0, 10.0)

# print("\nPSO completed\n")
# print("\nBest solution found:")
# print(["%.6f"%best_position[k] for k in range(dim)])
# fitnessVal = fitness(best_position)
# print("fitness of best solution = %.6f" % fitnessVal)

# print("\nEnd particle swarm for sphere function\n")


# dim = 13
# fitness = objective
# cons_min = [0.05, 15000, 50, 0.01, 100, 1.5, 0, 0.2, 0.05, 0.1, 0, 0.00001, 0.01]
# cons_max = [10, 22000, 100, 0.09, 150, 2.5, 0, 0.8, 0.25, 0.9, 0, 0.00009, 0.09]
# num_particles = 20
# max_iter = 20
# best_position = pso(fitness, max_iter, num_particles, dim, cons_min, cons_max)
# print("\nPSO completed\n")
# print("\nBest solution found:")
# print(["%.6f" % best_position[k] for k in range(dim)])
#
# fitness_val = fitness(best_position)
# print("fitness of best solution = %.6f" % fitness_val)
#
# print("\nEnd particle swarm for sphere function\n")
